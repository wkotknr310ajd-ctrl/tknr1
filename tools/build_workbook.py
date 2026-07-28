# -*- coding: utf-8 -*-
"""
勤務変更申請・承認システム のテンプレートブック(.xlsx)を生成するスクリプト。

生成されたテンプレートは以下の手順で完成させる(詳細は docs/導入手順.md 参照):
  1. このスクリプトで output/勤務変更管理システム.xlsx を生成
  2. Excelで開き、名前を付けて保存 → ファイルの種類を「Excel マクロ有効ブック(*.xlsm)」に変更
  3. Alt+F11 でVBAエディタを開き、vba/ 配下の .bas ファイルをすべてインポート
  4. ThisWorkbook のコード欄に vba/ThisWorkbook.txt の内容を貼り付け
  5. 職員マスタへの登録(RegisterStaffマクロ)を行い、共有フォルダへ配置
"""
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

TARGET_MONTH = datetime.date(2026, 8, 1)  # サンプルの対象年月(初回セットアップ時に変更してください)

SAMPLE_STAFF = ["サンプル職員1", "サンプル職員2", "サンプル職員3", "サンプル職員4", "サンプル職員5"]
SAMPLE_SHIFTS = ["日勤", "×", "夜A", "夜B", "夜C", "早", "遅", "年", "○", "研", ""]

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
LOCK_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()
wb.remove(wb.active)


def style_input_cell(cell):
    cell.fill = INPUT_FILL
    cell.border = BORDER
    cell.protection = openpyxl.styles.Protection(locked=False)


def style_label_cell(cell):
    cell.font = LABEL_FONT


# ------------------------------------------------------------------
# 1. シフト表
# ------------------------------------------------------------------
ws = wb.create_sheet("シフト表")
ws.sheet_view.showGridLines = False

ws["A1"] = "シフト表"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:E1")
ws["G1"] = "対象年月(設定シートB1で変更)"
ws["G1"].font = Font(italic=True, size=9, color="808080")

ws["A3"] = "職員氏名"
ws["A3"].fill = HEADER_FILL
ws["A3"].font = HEADER_FONT
ws["A4"] = "曜日"
ws["A4"].fill = HEADER_FILL
ws["A4"].font = HEADER_FONT

for d in range(1, 32):
    col = get_column_letter(1 + d)  # B=1日 ... AF=31日
    c3 = ws[f"{col}3"]
    c3.value = d
    c3.fill = HEADER_FILL
    c3.font = HEADER_FONT
    c3.alignment = Alignment(horizontal="center")
    c4 = ws[f"{col}4"]
    c4.value = (
        f'=IF({col}$3<=DAY(EOMONTH(設定!$B$1,0)),'
        f'TEXT(DATE(YEAR(設定!$B$1),MONTH(設定!$B$1),{col}$3),"aaa"),"")'
    )
    c4.alignment = Alignment(horizontal="center")
    ws.column_dimensions[col].width = 5

ws.column_dimensions["A"].width = 16

for i, name in enumerate(SAMPLE_STAFF):
    row = 5 + i
    ws.cell(row=row, column=1, value=name)
    for d in range(1, 32):
        col = 1 + d
        val = SAMPLE_SHIFTS[(i + d) % len(SAMPLE_SHIFTS)] if d <= 20 else ""
        ws.cell(row=row, column=col, value=val).alignment = Alignment(horizontal="center")

# シフト表全体をロックしてシート保護(VBAのUserInterfaceOnly保護で上書き運用)
for row in ws.iter_rows(min_row=1, max_row=60, min_col=1, max_col=33):
    for cell in row:
        cell.protection = openpyxl.styles.Protection(locked=True)
ws.protection.sheet = True
ws.protection.password = "shift-sys-2026"
ws.freeze_panes = "B5"

# ------------------------------------------------------------------
# 2. 申請
# ------------------------------------------------------------------
ws = wb.create_sheet("申請")
ws["A1"] = "勤務変更申請"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:D1")

fields = [
    (3, "申請者氏名", ""),
    (4, "パスワード", ""),
    (5, "対象日", ""),
    (6, "現在の勤務(自動表示)", "=IFERROR(INDEX(シフト表!$B$5:$AF$200,MATCH($B$3,シフト表!$A$5:$A$200,0),"
                          "MATCH(DAY($B$5),シフト表!$B$3:$AF$3,0)),\"-\")"),
    (7, "変更後の勤務", ""),
    (8, "変更理由(任意)", ""),
]
for row, label, formula in fields:
    lc = ws.cell(row=row, column=1, value=label)
    style_label_cell(lc)
    ic = ws.cell(row=row, column=2, value=formula if formula else None)
    if formula:
        ic.fill = LOCK_FILL
        ic.protection = openpyxl.styles.Protection(locked=True)
    else:
        style_input_cell(ic)
    ic.border = BORDER

ws["B5"].number_format = "yyyy/mm/dd"
ws["A10"] = "入力後、下の「申請する」ボタンを押してください。(ボタンは初回に自動作成されます)"
ws["A10"].font = Font(italic=True, size=9, color="808080")
ws.merge_cells("A10:F10")

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 28

ws.protection.sheet = True
ws.protection.password = "shift-sys-2026"

# ------------------------------------------------------------------
# 3. 承認
# ------------------------------------------------------------------
ws = wb.create_sheet("承認")
ws["A1"] = "勤務変更 承認"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:D1")

fields = [
    (3, "対象申請ID", "承認待ち一覧からIDを選択"),
    (4, "承認者氏名", ""),
    (5, "パスワード", ""),
    (6, "判定(承認/却下)", ""),
]
for row, label, note in fields:
    lc = ws.cell(row=row, column=1, value=label)
    style_label_cell(lc)
    ic = ws.cell(row=row, column=2)
    style_input_cell(ic)
    ic.border = BORDER
    if note:
        nc = ws.cell(row=row, column=4, value=note)
        nc.font = Font(italic=True, size=9, color="808080")

ws["A11"] = "▼ 承認待ち一覧(自動更新)"
ws["A11"].font = Font(bold=True)
headers = ["申請ID", "申請者", "対象日", "変更前", "変更後", "変更理由", "申請日時"]
for i, h in enumerate(headers):
    c = ws.cell(row=12, column=1 + i, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT

for i in range(1, 8):
    ws.column_dimensions[get_column_letter(i)].width = 16

dv_decision = DataValidation(type="list", formula1='"承認,却下"', allow_blank=True, showErrorMessage=True)
ws.add_data_validation(dv_decision)
dv_decision.add(ws["B6"])

ws.protection.sheet = False  # 入力セルが多いため保護なし(パスワード判定はマクロ側で実施)

# ------------------------------------------------------------------
# 4. ロールバック
# ------------------------------------------------------------------
ws = wb.create_sheet("ロールバック")
ws["A1"] = "承認済み変更の取り消し(ロールバック)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:D1")

fields = [
    (3, "対象申請ID", "承認済み一覧からIDを選択"),
    (4, "承認者氏名", "取り消しにも承認権限者のパスワードが必要です"),
    (5, "パスワード", ""),
]
for row, label, note in fields:
    lc = ws.cell(row=row, column=1, value=label)
    style_label_cell(lc)
    ic = ws.cell(row=row, column=2)
    style_input_cell(ic)
    ic.border = BORDER
    if note:
        nc = ws.cell(row=row, column=4, value=note)
        nc.font = Font(italic=True, size=9, color="808080")

ws["A10"] = "▼ 承認済み一覧(自動更新)"
ws["A10"].font = Font(bold=True)
headers = ["申請ID", "申請者", "対象日", "変更前", "変更後(現在値)", "承認者", "承認日時"]
for i, h in enumerate(headers):
    c = ws.cell(row=11, column=1 + i, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT

for i in range(1, 8):
    ws.column_dimensions[get_column_letter(i)].width = 16

ws.protection.sheet = False

# ------------------------------------------------------------------
# 5. 履歴
# ------------------------------------------------------------------
ws = wb.create_sheet("履歴")
headers = ["申請ID", "申請日時", "申請者", "対象日", "変更前", "変更後", "変更理由",
           "ステータス", "承認者", "承認/処理日時", "元申請ID(ロールバック用)"]
for i, h in enumerate(headers):
    c = ws.cell(row=1, column=1 + i, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    ws.column_dimensions[get_column_letter(1 + i)].width = 18

ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:K1"

for row in ws.iter_rows(min_row=1, max_row=2000, min_col=1, max_col=11):
    for cell in row:
        cell.protection = openpyxl.styles.Protection(locked=True)
ws.protection.sheet = True
ws.protection.password = "shift-sys-2026"
ws.protection.autoFilter = False

# ------------------------------------------------------------------
# 6. 職員マスタ (非表示)
# ------------------------------------------------------------------
ws = wb.create_sheet("職員マスタ")
headers = ["氏名", "権限(一般/一般・承認者)", "ソルト", "パスワードハッシュ", "登録日時"]
for i, h in enumerate(headers):
    c = ws.cell(row=1, column=1 + i, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    ws.column_dimensions[get_column_letter(1 + i)].width = 24
ws.sheet_state = "veryHidden"

# ------------------------------------------------------------------
# 7. 設定 (非表示)
# ------------------------------------------------------------------
ws = wb.create_sheet("設定")
ws["A1"] = "対象年月初日"
ws["B1"] = TARGET_MONTH
ws["B1"].number_format = "yyyy/mm/dd"
ws["A2"] = "次回申請ID通番"
ws["B2"] = 0
ws["A3"] = "メモ"
ws["B3"] = "シート保護パスワードは vba/modCommon.bas の SHEET_PROTECT_PASSWORD 定数で管理(運用上の誤操作防止用。真の権限管理は職員マスタのハッシュ照合で行う)"
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 60
ws.sheet_state = "veryHidden"

wb.active = 1  # 申請シートを既定表示に
out_path = "/home/user/tknr1/output/勤務変更管理システム.xlsx"
wb.save(out_path)
print("saved:", out_path)
